# -*-*-*-*- coding: utf-8 -*-*-*-*-
# @Time    : 2019/12/26 下午2:28        
# @Author  : LpL                    
# @Email   : peilun2050@gmail.com    
# -*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*
import os
from openpyxl import Workbook, load_workbook

from slowworld.conf.config import INCHANNEL_PATH


class ExcelUtil:
    @staticmethod
    def read_xlsx(filename, start, indexs):
        # 从filename文件的第一个sheet的第start行开始读取indexs这些列
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        ws = wb[sheet_names[0]]
        number = ws.max_row
        data_list = list()
        for i in range(start, number + 1):
            line = list()
            for j in indexs:
                cel = ws.cell(row=i, column=j).value
                line.append(cel)
            data_list.append(line)
        return data_list

    @staticmethod
    def write_xlsx(filename, start, indexs, data):
        # 创建filename文件，并从start行开始把数据写到indexs列中
        wb = Workbook()
        ws = wb.active
        number = len(data or [])
        num = 0
        for i in range(start, number + 1):
            for j in range(len(indexs)):
                ws.cell(row=i, column=indexs[j]).value = data[num][j]
            num += 1
        wb.save(filename)

    @staticmethod
    def load_xlsx(filename, start=2):
        # 加载整个表单
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        ws = wb[sheet_names[0]]
        number = ws.max_row
        data = list()
        titles = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for r in range(start, number + 1):
            line = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if any(line):
                data.append(line)
        return titles, data

    @staticmethod
    def get_titles(filename):
        # 获取标题
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        ws = wb[sheet_names[0]]
        titles = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        return titles


if __name__ == '__main__':
    # ExcelUtil.read_xlsx(os.path.join(INCHANNEL_PATH, "aaaa.xlsx"), 2, [1,2])
    # data = [[1,"a"],[2,"b"],[3,"c"],[4,"d"]]
    # ExcelUtil.write_xlsx(os.path.join(INCHANNEL_PATH, "bbb.xlsx"),2,[2,3],data)
    # ExcelUtil.load_xlsx(os.path.join(INCHANNEL_PATH, "cecece.xlsx"))
    ExcelUtil.get_titles("../../data/model_input/titles:1:157795115758.xlsx")
